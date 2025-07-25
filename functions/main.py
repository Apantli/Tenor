# Welcome to Cloud Functions for Firebase for Python!
# To get started, simply uncomment the below code or create your own.
# Deploy with `firebase deploy`

from firebase_functions import https_fn
from firebase_admin import initialize_app
import math
import time
import fitz
from flask import request, jsonify
from docx import Document
from openpyxl import load_workbook

initialize_app()

@https_fn.on_request()
def analyze_emotion(req: https_fn.Request) -> https_fn.Response:
    return {"emotion": "Happy", "timestamp": math.floor(time.time() * 1000)}

@https_fn.on_request()
def parse_file(req: https_fn.Request) -> https_fn.Response:
    if 'file' not in req.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    processed_files = []
    
    for file in req.files.getlist('file'):
        try:
            extension = file.filename.split('.')[-1].lower()
            
            if extension == "pdf":
                processed_files.append({"name": file.filename, "text": parse_pdf(file)})
            elif extension == "docx":
                processed_files.append({"name": file.filename, "text": extract_text_from_docx(file)})
            elif extension == "xlsx":
                processed_files.append({"name": file.filename, "text": extract_text_from_xlsx(file)})
            else:
                raise ValueError("Unsupported file type")
            
        except Exception as e:
            print(f"Error processing file {file}: {e}")
            pass

    return jsonify({"data": processed_files}), 200


def parse_pdf(pdf_file) -> https_fn.Response:
    try:
        doc = fitz.open(stream=pdf_file.read(), filetype="pdf")
        text = "\n".join(page.get_text("text") for page in doc)
        return text
    except Exception as e:
        print(f"Error parsing PDF: {e}")
        raise e

def extract_text_from_docx(file):
    doc = Document(file)
    text = "\n".join([para.text for para in doc.paragraphs])
    return text

def extract_text_from_xlsx(file):
    wb = load_workbook(file, data_only=True)
    extracted_text = []
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            extracted_text.append(" ".join(str(cell) for cell in row if cell is not None))
    
    return "\n".join(extracted_text)